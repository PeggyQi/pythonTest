import xlrd
#打开文件，获取excel文件的workbook（工作簿）对象,从1个sheet中读取数据写入新建的sheet中
import openpyxl

workbook = openpyxl.load_workbook(r"C:\Users\Administrator\Desktop\新发行债券-公司债.xlsx")
sheet=workbook['万得']
# 创建一个sheet对象，一个sheet对象对应Excel文件中的一张表格。
sheet2 = workbook.create_sheet()
sheet2.title = "test"
for i in range(1,sheet.max_row):
    for j in range(1,12):
        sheet2.cell(i,j).value=sheet.cell(i,j).value
workbook.save(r"C:\Users\Administrator\Desktop\新发行债券-公司债.xlsx")
# 其中的test是这张表的名字,cell_overwrite_ok，表示是否可以覆盖单元格，其实是Worksheet实例化的一个参数，默认值是False
